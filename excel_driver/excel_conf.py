"""
Excel的配置文件
"""
from openpyxl.styles import PatternFill,Font


def write(cell,row,column,status):
    """
    status: 1为通过
            其他不通过
    """
    if status==1:
        cell(row=row,column=column).value='Pass'
        # 绿色+加粗
        cell(row=row,column=column).fill=PatternFill('solid',fgColor='AACF91')
        cell(row=row,column=column).font=Font(bold=True)
    else:
        cell(row=row, column=column).value = 'Failed'
        # 红色+加粗
        cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
        cell(row=row, column=column).font = Font(bold=True)